"""
Normalize grade data from Excel sheets into a structured format.
Extracts student data from complex layouts and consolidates into single rows.
"""
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple

# File paths
INPUT_FILE = 'data/Copy of Copy 4th  2024_25 LA Benchmark.xlsx'
OUTPUT_FILE = 'data/normalized_grades.xlsx'

# Grade level mapping
GRADE_LEVEL_MAP = {
    'K 2021': 'Kindergarten',
    'First 2122': 'First',
    'Second 2223': 'Second',
    'Third 2324': 'Third',
    'Fourth 2425': 'Fourth'
}


def is_valid_student_name(name: str) -> bool:
    """Check if a name is a valid student name (not a section header)"""
    if pd.isna(name) or name == '':
        return False
    
    name_str = str(name).strip()
    
    # Skip section headers (like "KM", "1N", "2A", "3KB", "4R", "1R", "3NB")
    if re.match(r'^\d+[A-Z]+$', name_str) or name_str in ['KM', '1N', '2A', '3KB', '4R', '1R', '3NB']:
        return False
    
    # Must have at least 2 characters and start with a letter
    if len(name_str) < 2 or not name_str[0].isalpha():
        return False
    
    return True


def normalize_reading_level(level: str) -> Optional[str]:
    """Normalize reading level format (e.g., 'C+', 'C-', 'C/D' -> standardized format)"""
    if pd.isna(level) or level == '':
        return None
    
    level_str = str(level).strip().upper()
    
    # Handle ranges like "C/D", "P/Q", "M/N", "N/O"
    if '/' in level_str:
        parts = level_str.split('/')
        # Take the first level as primary, note range in original
        return parts[0].strip()
    
    # Handle plus/minus (C+, C-, etc.) - keep the base letter
    # Remove + and - but preserve the letter
    level_str = re.sub(r'[+\-]', '', level_str)
    
    # Extract just letters and numbers (for levels like "aa", "A", "1")
    level_str = re.sub(r'[^A-Z0-9]', '', level_str)
    
    return level_str if level_str else None


def normalize_grade_value(value: str) -> Optional[float]:
    """Convert various grade formats to numeric values"""
    if pd.isna(value) or value == '':
        return None
    
    value_str = str(value).strip()
    
    # Handle letter grades
    letter_grade_map = {
        'A': 4.0, 'B': 3.0, 'C': 2.0, 'D': 1.0, 'F': 0.0,
        'H': 4.0, 'OK': 2.5, 'M': 1.5, 'VLX': 2.0, 'P': 0.5
    }
    
    value_upper = value_str.upper()
    if value_upper in letter_grade_map:
        return letter_grade_map[value_upper]
    
    # Handle percentages (e.g., "83", "100+")
    if value_str.replace('+', '').replace('-', '').isdigit():
        num_val = float(value_str.replace('+', '').replace('-', ''))
        # Normalize to 0-4 scale if > 4 (assuming percentage)
        if num_val > 4:
            return num_val / 25.0  # Convert percentage to 0-4 scale
        return num_val
    
    # Handle fractions like "14/15"
    if '/' in value_str and value_str.replace('/', '').replace('-', '').isdigit():
        parts = value_str.split('/')
        if len(parts) == 2 and parts[1] != '0':
            return float(parts[0]) / float(parts[1]) * 4.0
    
    # Handle dates (convert to None for now)
    if re.match(r'\d{4}-\d{2}-\d{2}', value_str):
        return None
    
    return None


def extract_kindergarten_data(df: pd.DataFrame) -> pd.DataFrame:
    """Extract and normalize kindergarten data"""
    students = []
    
    # Find where student data starts (skip header rows 0-3)
    for idx in range(4, len(df)):
        student_name = df.iloc[idx, 0]
        
        if not is_valid_student_name(student_name):
            continue
        
        student_name = str(student_name).strip()
        
        student_data = {
            'Student_Name': student_name,
            'Grade_Level': 'Kindergarten',
            'Reading_Level_Fall': normalize_reading_level(df.iloc[idx, 2]) if idx < len(df) else None,
            'Reading_Level_Winter': normalize_reading_level(df.iloc[idx, 3]) if idx < len(df) else None,
            'Reading_Level_Spring': normalize_reading_level(df.iloc[idx, 4]) if idx < len(df) else None,
            'Reading_Level_EOY': normalize_reading_level(df.iloc[idx, 5]) if idx < len(df) else None,
            'Sight_Words_SeptNov': df.iloc[idx, 6] if idx < len(df) else None,
            'Sight_Words_Winter': df.iloc[idx, 7] if idx < len(df) else None,
            'Sight_Words_Spring': df.iloc[idx, 8] if idx < len(df) else None,
            'Sight_Words_EOY': df.iloc[idx, 9] if idx < len(df) else None,
            'Alphabet_Naming': df.iloc[idx, 11] if idx < len(df) else None,
            'PAR_Fall': df.iloc[idx, 13] if idx < len(df) else None,
            'PAR_EOY': df.iloc[idx, 14] if idx < len(df) else None,
            'Concerns': df.iloc[idx, 18] if idx < len(df) and len(df.columns) > 18 else None,
        }
        
        # Store original values
        student_data['Reading_Level_Fall_Original'] = df.iloc[idx, 2] if idx < len(df) else None
        student_data['Reading_Level_Winter_Original'] = df.iloc[idx, 3] if idx < len(df) else None
        student_data['Reading_Level_Spring_Original'] = df.iloc[idx, 4] if idx < len(df) else None
        student_data['Reading_Level_EOY_Original'] = df.iloc[idx, 5] if idx < len(df) else None
        
        students.append(student_data)
    
    return pd.DataFrame(students)


def extract_first_grade_data(df: pd.DataFrame) -> pd.DataFrame:
    """Extract and normalize first grade data"""
    students = []
    
    for idx in range(4, len(df)):
        student_name = df.iloc[idx, 0]
        
        if not is_valid_student_name(student_name):
            continue
        
        student_name = str(student_name).strip()
        
        student_data = {
            'Student_Name': student_name,
            'Grade_Level': 'First',
            'Reading_Level_Fall': normalize_reading_level(df.iloc[idx, 2]) if idx < len(df) else None,
            'Reading_Level_Winter': normalize_reading_level(df.iloc[idx, 3]) if idx < len(df) else None,
            'Reading_Level_Spring': normalize_reading_level(df.iloc[idx, 4]) if idx < len(df) else None,
            'Reading_Level_EOY': normalize_reading_level(df.iloc[idx, 5]) if idx < len(df) else None,
            'Sight_Words_SeptNov': df.iloc[idx, 6] if idx < len(df) else None,
            'Sight_Words_Winter': df.iloc[idx, 7] if idx < len(df) else None,
            'Sight_Words_Spring': df.iloc[idx, 8] if idx < len(df) else None,
            'Sight_Words_EOY': df.iloc[idx, 9] if idx < len(df) else None,
            'Spelling_Fall': df.iloc[idx, 11] if idx < len(df) else None,
            'Benchmark_Fall': df.iloc[idx, 12] if idx < len(df) else None,
            'Benchmark_Spring': df.iloc[idx, 13] if idx < len(df) else None,
            'PAR_Fall': df.iloc[idx, 14] if idx < len(df) else None,
            'PAR_EOY': df.iloc[idx, 15] if idx < len(df) else None,
            'Concerns': df.iloc[idx, 20] if idx < len(df) and len(df.columns) > 20 else None,
        }
        
        # Store original values
        student_data['Reading_Level_Fall_Original'] = df.iloc[idx, 2] if idx < len(df) else None
        student_data['Reading_Level_Winter_Original'] = df.iloc[idx, 3] if idx < len(df) else None
        student_data['Reading_Level_Spring_Original'] = df.iloc[idx, 4] if idx < len(df) else None
        student_data['Reading_Level_EOY_Original'] = df.iloc[idx, 5] if idx < len(df) else None
        
        students.append(student_data)
    
    return pd.DataFrame(students)


def extract_second_grade_data(df: pd.DataFrame) -> pd.DataFrame:
    """Extract and normalize second grade data"""
    students = []
    
    for idx in range(4, len(df)):
        student_name = df.iloc[idx, 0]
        
        if not is_valid_student_name(student_name):
            continue
        
        student_name = str(student_name).strip()
        
        student_data = {
            'Student_Name': student_name,
            'Grade_Level': 'Second',
            'Reading_Level_1EOY': normalize_reading_level(df.iloc[idx, 2]) if idx < len(df) else None,
            'Reading_Level_Fall': normalize_reading_level(df.iloc[idx, 3]) if idx < len(df) else None,
            'Reading_Level_Winter': normalize_reading_level(df.iloc[idx, 4]) if idx < len(df) else None,
            'Reading_Level_Spring': normalize_reading_level(df.iloc[idx, 5]) if idx < len(df) else None,
            'Reading_Level_EOY': normalize_reading_level(df.iloc[idx, 6]) if idx < len(df) else None,
            'Concerns': None,  # No concerns column in second grade sheet
        }
        
        # Store original values
        student_data['Reading_Level_1EOY_Original'] = df.iloc[idx, 2] if idx < len(df) else None
        student_data['Reading_Level_Fall_Original'] = df.iloc[idx, 3] if idx < len(df) else None
        student_data['Reading_Level_Winter_Original'] = df.iloc[idx, 4] if idx < len(df) else None
        student_data['Reading_Level_Spring_Original'] = df.iloc[idx, 5] if idx < len(df) else None
        student_data['Reading_Level_EOY_Original'] = df.iloc[idx, 6] if idx < len(df) else None
        
        students.append(student_data)
    
    return pd.DataFrame(students)


def extract_third_grade_data(df: pd.DataFrame) -> pd.DataFrame:
    """Extract and normalize third grade data"""
    students = []
    
    for idx in range(4, len(df)):
        student_name = df.iloc[idx, 0]
        
        if not is_valid_student_name(student_name):
            continue
        
        student_name = str(student_name).strip()
        
        student_data = {
            'Student_Name': student_name,
            'Grade_Level': 'Third',
            'Reading_Level_2EOY': normalize_reading_level(df.iloc[idx, 1]) if idx < len(df) else None,
            'Reading_Level_Fall': normalize_reading_level(df.iloc[idx, 2]) if idx < len(df) else None,
            'Reading_Level_Winter': normalize_reading_level(df.iloc[idx, 4]) if idx < len(df) else None,
            'Reading_Level_EOY': normalize_reading_level(df.iloc[idx, 6]) if idx < len(df) else None,
            'Spelling_Fall': df.iloc[idx, 3] if idx < len(df) else None,
            'Slingerlands_Fall': df.iloc[idx, 5] if idx < len(df) else None,
            'Spelling_EOY': df.iloc[idx, 7] if idx < len(df) else None,
            'Benchmark_Spring': df.iloc[idx, 8] if idx < len(df) else None,
            'Concerns': df.iloc[idx, 9] if idx < len(df) and len(df.columns) > 9 else None,
        }
        
        # Store original values
        student_data['Reading_Level_2EOY_Original'] = df.iloc[idx, 1] if idx < len(df) else None
        student_data['Reading_Level_Fall_Original'] = df.iloc[idx, 2] if idx < len(df) else None
        student_data['Reading_Level_Winter_Original'] = df.iloc[idx, 4] if idx < len(df) else None
        student_data['Reading_Level_EOY_Original'] = df.iloc[idx, 6] if idx < len(df) else None
        
        students.append(student_data)
    
    return pd.DataFrame(students)


def extract_fourth_grade_data(df: pd.DataFrame) -> pd.DataFrame:
    """Extract and normalize fourth grade data"""
    students = []
    
    for idx in range(4, len(df)):
        student_name = df.iloc[idx, 0]
        
        if not is_valid_student_name(student_name):
            continue
        
        student_name = str(student_name).strip()
        
        student_data = {
            'Student_Name': student_name,
            'Grade_Level': 'Fourth',
            'Reading_Level_3EOY': normalize_reading_level(df.iloc[idx, 1]) if idx < len(df) else None,
            'Reading_Level_Fall': normalize_reading_level(df.iloc[idx, 2]) if idx < len(df) else None,
            'Reading_Level_Winter': normalize_reading_level(df.iloc[idx, 4]) if idx < len(df) else None,
            'Reading_Level_EOY': normalize_reading_level(df.iloc[idx, 5]) if idx < len(df) else None,
            'Spelling_Fall': df.iloc[idx, 3] if idx < len(df) else None,
            'Spelling_Spring': df.iloc[idx, 6] if idx < len(df) else None,
            'Concerns': df.iloc[idx, 7] if idx < len(df) and len(df.columns) > 7 else None,
        }
        
        # Store original values
        student_data['Reading_Level_3EOY_Original'] = df.iloc[idx, 1] if idx < len(df) else None
        student_data['Reading_Level_Fall_Original'] = df.iloc[idx, 2] if idx < len(df) else None
        student_data['Reading_Level_Winter_Original'] = df.iloc[idx, 4] if idx < len(df) else None
        student_data['Reading_Level_EOY_Original'] = df.iloc[idx, 5] if idx < len(df) else None
        
        students.append(student_data)
    
    return pd.DataFrame(students)


def combine_all_grades(all_dfs: List[pd.DataFrame]) -> pd.DataFrame:
    """Combine all grade DataFrames into one unified structure"""
    # Get all unique columns
    all_columns = set()
    for df in all_dfs:
        all_columns.update(df.columns)
    
    # Define preferred column order
    priority_columns = ['Student_Name', 'Grade_Level']
    other_columns = sorted([c for c in all_columns if c not in priority_columns])
    
    # Reorder: priority columns first, then others
    column_order = priority_columns + other_columns
    
    # Reindex all DataFrames to have the same columns
    combined_dfs = []
    for df in all_dfs:
        df_reindexed = df.reindex(columns=column_order)
        combined_dfs.append(df_reindexed)
    
    # Concatenate all DataFrames
    combined = pd.concat(combined_dfs, ignore_index=True)
    
    # Sort by student name first, then grade level (K, 1, 2, 3, 4)
    grade_order = ['Kindergarten', 'First', 'Second', 'Third', 'Fourth']
    combined['Grade_Order'] = combined['Grade_Level'].map({g: i for i, g in enumerate(grade_order)})
    combined = combined.sort_values(['Student_Name', 'Grade_Order']).drop('Grade_Order', axis=1)
    
    # Reset column order after sorting
    combined = combined[column_order]
    
    return combined


def main():
    """Main function to normalize grade data"""
    print("Loading Excel file...")
    wb = load_workbook(INPUT_FILE)
    sheet_names = wb.sheetnames
    
    print(f"Found sheets: {sheet_names}")
    
    all_student_data = []
    
    # Process each sheet
    for sheet_name in sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=None)
        
        if sheet_name == 'K 2021':
            student_df = extract_kindergarten_data(df)
        elif sheet_name == 'First 2122':
            student_df = extract_first_grade_data(df)
        elif sheet_name == 'Second 2223':
            student_df = extract_second_grade_data(df)
        elif sheet_name == 'Third 2324':
            student_df = extract_third_grade_data(df)
        elif sheet_name == 'Fourth 2425':
            student_df = extract_fourth_grade_data(df)
        else:
            print(f"  Unknown sheet format, skipping...")
            continue
        
        print(f"  Extracted {len(student_df)} students")
        all_student_data.append(student_df)
    
    # Combine all grades
    print("\nCombining all grade data...")
    combined_df = combine_all_grades(all_student_data)
    
    print(f"\nTotal students across all grades: {len(combined_df)}")
    print(f"Total columns: {len(combined_df.columns)}")
    
    # Save to Excel
    print(f"\nSaving to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name='All_Grades', index=False)
    
    print("Done!")
    print(f"\nOutput file: {OUTPUT_FILE}")
    print(f"Shape: {combined_df.shape}")
    print(f"\nFirst few rows:")
    print(combined_df.head().to_string())


if __name__ == '__main__':
    main()
